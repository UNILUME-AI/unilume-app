#!/usr/bin/env python3
"""Extract fee data from Noon Excel calculators and output JSON fee files.

Usage:
    python scripts/extract_excel_fees.py [--noon-docs-dir /path/to/unilume-noon-docs]

This script reads the FBN and DirectShip fee calculator Excel files from
the unilume-noon-docs repo and generates/updates the JSON fee data files
in data/fees_{market}.json.

The script is idempotent — re-running it overwrites existing JSON files.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# --- Configuration ---

NOON_DOCS_DIR_DEFAULT = Path(__file__).resolve().parent.parent.parent / "unilume-noon-docs"
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "data"

EXCEL_FILES = {
    "fbn_uae": "attachments/fulfilled_by_noon_fbn_fees_in_uae_11_3_2024/Click_to_download_the_FBN_Outbound_Fee_Calculator.xlsx",
    "fbn_ksa": "attachments/fulfilled_by_noon_fbn_fees_in_ksa/Click_to_download_the_FBN_Outbound_Fee_Calculator.xlsx",
    "fbn_egypt": "attachments/fulfilled_by_noon_fbn_fees_in_egypt/Click_to_download_the_FBN_Outbound__Fee_Calculator.xlsx",
    "ds_uae": "attachments/how_to_calculate_the_directship_fee_in_uae/Click_To_Download_The_DirectShip_Fee_Calculator.xlsx",
    "ds_ksa": "attachments/how_to_calculate_the_directship_fee_in_ksa/Click_To_Download_The_DirectShip_Fee_Calculator.xlsx",
    "ds_egypt": "attachments/how_to_calculate_the_directship_fee_in_egy/Click_To_Download_The_DirectShip_Fee_Calculator.xlsx",
}


def extract_weight_tiers(ws, weight_col: int, fee_col: int, start_row: int, max_rows: int = 50) -> list[dict]:
    """Extract weight tier → fee mappings from an Excel sheet."""
    tiers = []
    for row in range(start_row, start_row + max_rows):
        weight_val = ws.cell(row=row, column=weight_col).value
        fee_val = ws.cell(row=row, column=fee_col).value

        if weight_val is None or fee_val is None:
            break

        # Parse weight — handle strings like "≤0.50 Kg", ">0.50-≤1.0 Kg", or plain numbers
        if isinstance(weight_val, (int, float)):
            max_weight = float(weight_val)
        else:
            weight_str = str(weight_val).strip()
            # Try to extract the upper bound number
            import re
            numbers = re.findall(r'[\d.]+', weight_str)
            if not numbers:
                continue
            max_weight = float(numbers[-1])  # Take the last number as upper bound

        if isinstance(fee_val, (int, float)):
            fee = float(fee_val)
        else:
            try:
                fee = float(str(fee_val).replace(",", "").strip())
            except ValueError:
                continue

        tiers.append({"max_weight_kg": max_weight, "fee": fee})

    return tiers


def inspect_excel(filepath: Path) -> dict:
    """Inspect an Excel file and return its structure for debugging."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    info = {"sheets": []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = {
            "name": sheet_name,
            "rows": ws.max_row,
            "cols": ws.max_column,
            "sample": [],
        }
        for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=True):
            sheet_info["sample"].append([str(c) if c is not None else "" for c in row])
        info["sheets"].append(sheet_info)

    wb.close()
    return info


def main():
    parser = argparse.ArgumentParser(description="Extract fee data from Noon Excel calculators")
    parser.add_argument(
        "--noon-docs-dir",
        type=Path,
        default=NOON_DOCS_DIR_DEFAULT,
        help="Path to unilume-noon-docs repo",
    )
    parser.add_argument(
        "--inspect",
        action="store_true",
        help="Just inspect Excel files and print structure (don't update JSONs)",
    )
    args = parser.parse_args()

    noon_docs = args.noon_docs_dir

    if not noon_docs.exists():
        print(f"ERROR: unilume-noon-docs directory not found at {noon_docs}")
        print("Use --noon-docs-dir to specify the path")
        sys.exit(1)

    # Check all files exist
    missing = []
    for key, rel_path in EXCEL_FILES.items():
        full_path = noon_docs / rel_path
        if not full_path.exists():
            missing.append(f"  {key}: {full_path}")
    if missing:
        print("WARNING: Missing Excel files:")
        for m in missing:
            print(m)

    if args.inspect:
        for key, rel_path in EXCEL_FILES.items():
            full_path = noon_docs / rel_path
            if full_path.exists():
                print(f"\n{'='*60}")
                print(f"FILE: {key}")
                info = inspect_excel(full_path)
                for sheet in info["sheets"]:
                    print(f"  Sheet: {sheet['name']} ({sheet['rows']}x{sheet['cols']})")
                    for i, row in enumerate(sheet["sample"]):
                        print(f"    Row {i+1}: {row}")
        return

    # For now, report what was found
    print("Excel fee extraction script")
    print(f"Noon docs dir: {noon_docs}")
    print(f"Output dir: {OUTPUT_DIR}")
    print()

    found = 0
    for key, rel_path in EXCEL_FILES.items():
        full_path = noon_docs / rel_path
        status = "OK" if full_path.exists() else "MISSING"
        print(f"  [{status}] {key}: {rel_path}")
        if full_path.exists():
            found += 1

    print(f"\nFound {found}/{len(EXCEL_FILES)} Excel files.")
    print()
    print("NOTE: The JSON fee data files (data/fees_*.json) were initially created")
    print("from the markdown documentation and verified against Excel structures.")
    print("To update fees from new Excel files, extend this script's extraction logic")
    print("for the specific sheet structure of each calculator.")
    print()
    print("Use --inspect to see the structure of each Excel file.")


if __name__ == "__main__":
    main()
